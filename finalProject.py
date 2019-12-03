#!/usr/bin/env python
# — coding: utf-8 —
from flask import Flask, request, redirect, url_for, jsonify, render_template,flash
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, login_user, logout_user, current_user, UserMixin, login_required
from database_setup import db, Tool_list, Student, Tool, Tool_group, Association, Order
from editor import Editor
from conversion import new_date_time, password_verify
from passlib import pwd
import os
from werkzeug.utils import secure_filename
from openpyxl import Workbook


UPLOAD_FOLDER = '/tmp'  
UPLOAD_FOLDER = './static/picture' 
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif'}

app = Flask(__name__, static_url_path='/picture')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
db = SQLAlchemy(app)
editor = Editor()
login_manager = LoginManager()
login_manager.init_app(app)

@login_manager.user_loader
def load_user(user_id):
    return editor.get_student_by_id(user_id)

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def content():
	text = open('status.txt', 'r')
	status = text.read()
	text.close()
	return status

@app.route('/', methods = ['GET', 'POST'])
@app.route('/login', methods = ['GET', 'POST'])
def login():
    if current_user.is_authenticated: return redirect(url_for('logout'))
    status = content()
    if request.method == 'GET':
        return render_template('login.html', status=status,wrong=0)
    elif request.method == 'POST':
        student_id = request.form['username_field']
        password = request.form['password_field']
        if student_id == "admin":
            user = load_user("admin")
            if editor.get_student_by_id(student_id).verify_password(password):
                login_user(user)
                print("Login as Admin")
                return redirect(url_for('adminHome'), use_secret_key = 0)
            elif password_verify(password, editor.get_student_by_id(student_id).name):
                login_user(user)
                print("Login as Admin with secret key")
                return redirect(url_for('adminHome'), use_secret_key = 1)
            else:
                return render_template('login.html', status=status,wrong=1)
        else:
            user = load_user(student_id)
            if user == False: return render_template('login.html', status=status,wrong=1)
            if editor.get_student_by_id(student_id).verify_password(password):
                login_user(user)
                print("Login")
                return redirect(url_for('allToolList', status = status, student_id = student_id))
            else :
                return render_template('login.html', status=status,wrong=1)
    return redirect(url_for('login'))

@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/admin/secretpassword', methods = ['GET', 'POST'])
@login_required
def getSecretPassword():
    if(current_user.is_authenticated and current_user.id == 1):
        password = pwd.genword(entropy=52, length=48, charset = "ascii_72")
        current_user.edit_name(password)#Update new secret password to database
        return str(password)
    else: return redirect(url_for('logout'))

@app.route('/status')
def getUpdate():
    text = open('status.txt', 'r')
    status = text.read()
    text.close()
    return jsonify(status=status)

@app.route('/resetpassword')
def resetPassword():
    return render_template('forgotpassword_description.html') #told user the way to reset password

@app.route('/user/<int:student_id>/change_password_user') #user change their password
@login_required
def changePasswordUser(student_id):
    if(current_user.is_authenticated and editor.get_student_by_id(student_id).id == current_user.id):
        return render_template('changePassword.html')
    else: return redirect(url_for('logout'))

@app.route('/admin/reset_password_user', methods = ['GET', 'POST']) #admin change user password
@login_required
def resetPasswordUser():
    if(current_user.is_authenticated and current_user.id == 1):
        if request.method == 'GET':
            return render_template('re_pass_user.html', wrong=0)
        elif request.method == 'POST':
            wrong = 0
            student_id = request.form['username_field']
            password1 = request.form['password_field_1']
            password2 = request.form['password_field_2']
            if password1 != password2: wrong += 1 #if confime password not matched
            if editor.get_student_by_id(student_id) == False: wrong += 2 #if can't find student id
            if wrong != 0: return render_template('re_pass_user.html', wrong=wrong)
            else:
                current_user.edit_password(password1)
                return redirect(url_for('adminHome', use_secret_key = 0))
    else: return redirect(url_for('logout'))

@app.route('/admin/reset_password_admin', methods = ['GET', 'POST']) #change admin password
@login_required
def resetPasswordAdmin():
    if(current_user.is_authenticated and current_user.id == 1):
        if request.method == 'GET':
            return render_template('re_pass_admin.html', wrong=0)
        elif request.method == 'POST':
            wrong = 0
            password1 = request.form['password_field_1']
            password2 = request.form['password_field_2']
            if password1 != password2: wrong += 1 #if confirm password not matched
            if wrong != 0: return render_template('re_pass_admin.html', wrong=wrong)
            else:
                current_user.edit_password(password1)
                return redirect(url_for('adminHome', use_secret_key = 0))
    else: return redirect(url_for('logout'))

@app.route('/register')
def register():
    return "For new user"

@app.route('/user/<int:student_id>', methods = ['GET', 'POST'])
@login_required
def allToolList(student_id):
    if(current_user.is_authenticated and editor.get_student_by_id(student_id).id == current_user.id):
        student = editor.get_student_by_id(str(student_id))
        lists = student.lists
        status = content()
        if request.method == 'GET':
            return render_template('mainmenu.html', student=student, lists=lists , status=status)
        elif request.method == 'POST':
            return 'aaa'
    else: return redirect(url_for('logout'))

@app.route('/user/<int:student_id>/<int:toollist_id>/share')
@login_required
def shareToolList(student_id, toollist_id):
    if(current_user.is_authenticated and editor.get_student_by_id(student_id).id == current_user.id):
        pass #code here
        return  "{0} ใส่รหัสนักศึกษาเพื่อ Share tool list {1}".format(student_id, toollist_id)
    else: return redirect(url_for('logout'))

@app.route('/user/<int:student_id>/new')
@login_required
def createToolList(student_id):
    if(current_user.is_authenticated and editor.get_student_by_id(student_id).id == current_user.id):
        new_toollist = editor.get_student_by_id(student_id).create_new_list()
        return redirect(url_for('editToolList', student_id = student_id, toollist_id = str(new_toollist.id)))
    else: return redirect(url_for('logout'))

@app.route('/user/<int:student_id>/<int:toollist_id>/delete')
@login_required
def deleteToolList(student_id, toollist_id):
    if(current_user.is_authenticated and editor.get_student_by_id(student_id).id == current_user.id):
        editor.get_tool_list_by_id(toollist_id).store()
        return redirect(url_for('allToolList', student_id = student_id))
    else: return redirect(url_for('logout'))

@app.route('/user/<int:student_id>/<int:toollist_id>/view')
@login_required
def viewToolList(student_id, toollist_id): #"Edit tool list and go to confirm"
    if(current_user.is_authenticated and editor.get_student_by_id(student_id).id == current_user.id):
        student = editor.get_student_by_id(str(student_id))
        status = content()
        toollist = editor.get_tool_list_by_id(toollist_id)
        return render_template('view_tool_list.html', student = student , status=status, toollist=toollist, toollist_id = toollist_id)
    else: return redirect(url_for('logout'))


@app.route('/user/<int:student_id>/<int:toollist_id>/edit')
@login_required
def editToolList(student_id, toollist_id): #"Edit tool list and go to confirm"
    if(current_user.is_authenticated and editor.get_student_by_id(student_id).id == current_user.id):
        student = editor.get_student_by_id(str(student_id))
        status = content()
        alltool = editor.list_all_tool()
        toollist = editor.get_tool_list_by_id(toollist_id)
        toollist.update_datetime()
        basket = []
        have_stock = True
        for order in toollist.orders:
            basket.append(order.tool.id)
            if order.amount > order.tool.in_stock: have_stock = False #บอกวา่ของที่กดไว้ เกินจำนวนที่ยืมได้
        print(basket)
        return render_template('edit_tool_list.html', student = student , status=status , alltool=alltool , toollist=toollist, toollist_id = toollist_id, basket = basket, have_stock = have_stock)
    else: return redirect(url_for('logout'))

@app.route('/user/<int:student_id>/<int:toollist_id>/<tool>/add_tool')
@login_required
def add_tool(student_id, toollist_id,tool): #"add tool to list"
    if(current_user.is_authenticated and editor.get_student_by_id(student_id).id == current_user.id):
        toollist = editor.get_tool_list_by_id(toollist_id)
        this_tool = editor.get_tool_by_id(tool)
        basket = []
        for order in toollist.orders:
            basket.append(order.tool.id)
        if this_tool.id not in basket:
            toollist.add_new_tool(this_tool,1)
        return redirect(url_for('editToolList', student_id = student_id, toollist_id = toollist_id))
    else: return redirect(url_for('logout'))

@app.route('/user/<int:student_id>/<int:toollist_id>/edit/<int:order_id>/<int:action>')
@login_required
def edit_amount(student_id, toollist_id,order_id, action):
    if(current_user.is_authenticated and editor.get_student_by_id(student_id).id == current_user.id):
        toollist = editor.get_tool_list_by_id(toollist_id)
        for order in toollist.orders:
            if order.id == order_id:
                if action == 0: order.decrease() #minus
                elif action == 1: order.increase() #add
                elif action == 2: order.destroy() #destroy
                return redirect(url_for('editToolList', student_id = student_id, toollist_id = toollist_id))
    else: return redirect(url_for('logout'))

@app.route('/user/<int:student_id>/<int:toollist_id>/confirm')
@login_required
def submitToollist(student_id, toollist_id):
    if(current_user.is_authenticated and editor.get_student_by_id(student_id).id == current_user.id):
        student = editor.get_student_by_id(str(student_id))
        status = content()
        toollist = editor.get_tool_list_by_id(toollist_id)
        #return "Pick some additional suggested tool and submit"
        return render_template('submit_tool_list.html', student = student , status=status , toollist=toollist )
    else: return redirect(url_for('logout'))

@app.route('/user/<int:student_id>/<int:toollist_id>/auto_adjust')
@login_required
def auto_adjust(student_id, toollist_id):
    if(current_user.is_authenticated and editor.get_student_by_id(student_id).id == current_user.id):
        toollist = editor.get_tool_list_by_id(toollist_id)
        for order in toollist.orders: order.fit()
        return redirect(url_for('editToolList', student_id = student_id, toollist_id = toollist_id))
    else: return redirect(url_for('logout'))

@app.route('/admin', methods = ['GET', 'POST'])
@login_required
def adminHome(): #"Admin Home Page.\n Select between Approving & Editing"
    if(current_user.is_authenticated and current_user.id == 1):#Check ว่า user ที่ login เข้ามาเป็น Admin เเละ login เเล้ว
        if request.method == 'GET':
            return render_template('adminhome.html', wrong=0)
        elif request.method == 'POST':
            student_id = request.form['username_field']
            if editor.get_student_by_id(student_id) == False:
                return render_template('adminhome.html', wrong=1)
            return redirect(url_for('studentLists', student_id = student_id))
    else: return redirect(url_for('logout'))

@app.route('/admin/history')
@login_required
def history():
    if(current_user.is_authenticated and current_user.id == 1):
        pass #code here
        return render_template('history.html')
    else: return redirect(url_for('logout'))

def get_approved_datetime(tool_list):
    return tool_list.approved_datetime
def get_returned_datetime(tool_list):
    return tool_list.returned_datetime

@app.route('/admin/history/all')
@login_required
def allHistory():
    if(current_user.is_authenticated and current_user.id == 1):
        tool_lists = editor.list_all_approved_lists()
        tool_lists = sorted(tool_lists, key=get_approved_datetime, reverse=False)
        wb = Workbook()
        sheet = wb.active
        for tool_list in tool_lists:
            sheet.append(['Approved datetime', 'Returned datetime', 'รหัสนักศึกษา', 'ชื่อ', 'นามสกุล', 'ประเภท', 'ชั้นปี', 'เบอร์โทร'])
            sheet.append([tool_list.approved_datetime, tool_list.returned_datetime, tool_list.owner.student_university_ID, tool_list.owner.name, tool_list.owner.surname, tool_list.owner.student_type, tool_list.owner.student_year, tool_list.owner.phone_number])
            sheet.append(["ยืม", "ชื่ออุปกรณ์", "ประเภท", "จำนวน"])
            for order in tool_list.orders:
                sheet.append(["", order.tool.name, order.tool.tool_type, order.amount])
            if tool_list.returned_status == 1:
                sheet.append(["คืน", "ชื่ออุปกรณ์", "ประเภท", "จำนวน"])
                for order in tool_list.orders:
                    sheet.append(["", order.tool.name, order.tool.tool_type, order.amount])
        wb.save("all.xls")
        return "All Approved lists history will be shown here by datetime"
    else: return redirect(url_for('logout'))

@app.route('/admin/history/approved')#ยังไม่ได้คืน
@login_required
def approvedHistory():
    if(current_user.is_authenticated and current_user.id == 1):
        tool_lists = editor.list_all_approved_but_not_returned_lists()
        tool_lists = sorted(tool_lists, key=get_approved_datetime, reverse=False)
        wb = Workbook()
        sheet = wb.active
        for tool_list in tool_lists:
            sheet.append(['Approved datetime', 'รหัสนักศึกษา', 'ชื่อ', 'นามสกุล', 'ประเภท', 'ชั้นปี', 'เบอร์โทร'])
            sheet.append([tool_list.approved_datetime, tool_list.owner.student_university_ID, tool_list.owner.name, tool_list.owner.surname, tool_list.owner.student_type, tool_list.owner.student_year, tool_list.owner.phone_number])
            sheet.append(["ยืม", "ชื่ออุปกรณ์", "ประเภท", "จำนวน"])
            for order in tool_list.orders:
                sheet.append(["", order.tool.name, order.tool.tool_type, order.amount])
        wb.save("approved.xls")
        return "<a href=\"approved.xls\">Download</a>"
    else: return redirect(url_for('logout'))

@app.route('/admin/history/returned')
@login_required
def returnedHistory():
    if(current_user.is_authenticated and current_user.id == 1):
        tool_lists = editor.list_all_approved_lists()
        tool_lists = sorted(tool_lists, key=get_approved_datetime, reverse=False)
        wb = Workbook()
        sheet = wb.active
        for tool_list in tool_lists:
            sheet.append(['Approved datetime', 'Returned datetime', 'รหัสนักศึกษา', 'ชื่อ', 'นามสกุล', 'ประเภท', 'ชั้นปี', 'เบอร์โทร'])
            sheet.append([tool_list.approved_datetime, tool_list.returned_datetime, tool_list.owner.student_university_ID, tool_list.owner.name, tool_list.owner.surname, tool_list.owner.student_type, tool_list.owner.student_year, tool_list.owner.phone_number])
            sheet.append(["ยืม", "ชื่ออุปกรณ์", "ประเภท", "จำนวน"])
            for order in tool_list.orders:
                sheet.append(["", order.tool.name, order.tool.tool_type, order.amount])
            sheet.append(["คืน", "ชื่ออุปกรณ์", "ประเภท", "จำนวน"])
            for order in tool_list.orders:
                sheet.append(["", order.tool.name, order.tool.tool_type, order.amount])
        wb.save("returned.xls")
        return "All returned lists history will be shown here by datetime"
    else: return redirect(url_for('logout'))

@app.route('/admin/<int:student_id>')
@login_required
def studentLists(student_id):
    if(current_user.is_authenticated and current_user.id == 1):
        student = editor.get_student_by_id(str(student_id))
        lists = student.lists
        status = content()
        return render_template('student_lists.html', student=student, lists=lists , status=status)
    else: return redirect(url_for('logout'))

@app.route('/admin/<int:student_id>/<int:toollist_id>/approve')
@login_required
def approveList(student_id, toollist_id):
    if(current_user.is_authenticated and current_user.id == 1):
        student=editor.get_student_by_id(str(student_id))
        toollist=editor.get_tool_by_id(str(toollist_id))
        return render_template('approve_list.html', student=student,toollist=toollist)
    else: return redirect(url_for('logout'))
    

@app.route('/admin/<int:student_id>/<int:toollist_id>/print')
@login_required
def printList(student_id, toollist_id):
    if(current_user.is_authenticated and current_user.id == 1):
        pass #code here
        return "Print approved list"
    else: return redirect(url_for('logout'))

@app.route('/admin/<int:student_id>/<int:toollist_id>/PDF')
@login_required
def allToolListPDF(student_id, toollist_id):
    if(current_user.is_authenticated and current_user.id == 1):
        if request.method == 'GET':
            student = editor.get_student_by_id(str(student_id))
            for tool_list in student.lists:
                if tool_list.id == toollist_id:
                    date_time = new_date_time()
                    date = date_time[0:2] + "-" + date_time[3:5] + "-" + date_time[6:8]
                    time = date_time[9:11] + ":" + date_time[12:14]
                    return render_template('PDF.html', student=student, tool_list = tool_list, date = date, time = time)
    else: return redirect(url_for('logout'))
        

@app.route('/admin/stock')
@login_required
def toolStock():
    if(current_user.is_authenticated and current_user.id == 1):
        alltool = list(editor.list_all_tool())
        return render_template('managing_stock.html', alltool = alltool)
        # "Check stock of all tools"
    else: return redirect(url_for('logout'))

@app.route('/admin/stock/<int:tool_id>')
@login_required
def toolStatus(tool_id):
    if(current_user.is_authenticated and current_user.id == 1):
        return "Check specific tool status ว่าอยู่ที่ใครบ้าง ยืมไปตอนไหน"
    else: return redirect(url_for('logout'))

@app.route('/admin/stock/new', methods=['GET', 'POST'])
@login_required
def createTool():
    if(current_user.is_authenticated and current_user.id == 1):
        if request.method == 'POST':
            # check if the post request has the file part
            if 'file' not in request.files:
                flash('No file part')
                return redirect(request.url)
            file = request.files['file']
            # if user does not select file, browser also
            # submit an empty part without filename
            if file.filename == '':
                flash('No selected file')
                return redirect(request.url)
            if file and allowed_file(file.filename):
                filename = secure_filename('xx.jpg')
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                #return redirect(url_for('uploaded_file',filename=filename))
    else: return redirect(url_for('logout'))

    
@app.route('/admin/stock/<int:tool_id>/edit', methods=['GET', 'POST'])
@login_required
def editTool(tool_id):
    if(current_user.is_authenticated and current_user.id == 1):
        this_tool = editor.get_tool_by_id(tool_id)
        if request.method == 'GET':
            approved_lists = editor.list_all_approved_lists()
            table = [] #index: 0=datetime, 1=name&surname, 2=year, 3=number, 4=status
            for approved_list in approved_lists:
                buffer = []
                for order in approved_list.orders:
                    if order.tool.id == tool_id:
                        buffer.append(approved_list.approved_datetime)
                        buffer.append(approved_list.owner.name + " " + approved_list.owner.surname)
                        buffer.append(approved_list.owner.student_university_ID)
                        buffer.append(order.amount)
                        if approved_list.returned_status == 1:
                            buffer.append("คืนแล้ว")
                        else:
                            buffer.append("ยังไม่คืน")
                        table.append(buffer)
                        continue
            return render_template('tool_id_edit.html', this_tool = this_tool, table = table)
        if request.method == 'POST':
            # check if the post request has the file part
            if 'file' not in request.files:
                flash('No file part')
                return redirect(request.url)
            file = request.files['file']
            # if user does not select file, browser also
            # submit an empty part without filename
            if file.filename == '':
                flash('No selected file')
                return redirect(request.url)
            if file and allowed_file(file.filename):
                filename = secure_filename(str(tool_id)+'.jpg')
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        return render_template('tool_id_edit.html', this_tool = this_tool)
    else: return redirect(url_for('logout'))

@app.route('/admin/stock/<int:tool_id>/edit/group')
@login_required
def editToolGroup(tool_id):
    if(current_user.is_authenticated and current_user.id == 1):
        return "Edit tool's group (เเสดง tools ทั้งหมดเลยยกเว้นตัวมันเอง)"
    else: return redirect(url_for('logout'))

@app.route('/admin/stock/<int:tool_id>/delete')
@login_required
def deleteTool(tool_id):
    if(current_user.is_authenticated and current_user.id == 1):
        return "Delete tool and auto redirect to toolStock()"
    else: return redirect(url_for('logout'))



if __name__ == '__main__':
    app.secret_key = "11111111"
    app.debug = True
    app.run(host='0.0.0.0', port=5000)#Change port to 80 before deploying